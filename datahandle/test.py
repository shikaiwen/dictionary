# coding:utf-8
'''
Created on 2017/12/20

@author: shikw
'''
import sys
import unittest
import openpyxl
from openpyxl.xml.constants import MAX_ROW
from builtins import str
import re
from _functools import reduce
import db 

print(__name__)
print(__file__)
# map() reduce() filter()

# >>> foo = [2, 18, 9, 22, 17, 24, 8, 12, 27]
# >>> 
# >>> print filter(lambda x: x % 3 == 0, foo)
# [18, 9, 24, 12, 27]
# >>> 
# >>> print map(lambda x: x * 2 + 10, foo)
# [14, 46, 28, 54, 44, 58, 26, 34, 64]
# >>> 
# >>> print reduce(lambda x, y: x + y, foo)
# 139

# reduce 
d = [3,6,2,45,5,3,7,7,5,66,43,886,43,8,54,65,78,55,33,55,76,58]
print(reduce(lambda x,y : x if x>y else y ,d ))

exit()
number_mapping = {'1': 'one',
                  '2': 'two',
                  '3': 'three'}
s = "1 testing 2 3"
print(re.sub(r'\d', lambda matchObj: number_mapping[matchObj.group()], s))


pattern = """
{ field: 'recid', caption: '&nbsp;', size: '1%',  resizable: true, sortable: true, render: function (record, index, column_index) {
                return '<div><input type="checkbox" value="1" recid="' + index + '"</div>';
                
            } },
            { field: 'index_4', caption: '約定日時', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[4]
            }},
            { field: 'index_14', caption: '注文手法（注文区分）', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return namepair[record[14]]
            }},
            { field: 'index_0', caption: '通貨ペアコード', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[0]
            }},
            { field: 'index_5', caption: '売買', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                //売買区分
                return record[5]== "0" ? "売":"買";
            }},
            { field: 'index_6', caption: '新規決済', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                //新規決済区分
                return record[6]== 0? "新規":"決済";
            }},
            { field: 'index_8', caption: '約定数量', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                //数量
                return FormatUtil.formatInteger(record[8])
            }},
            { field: 'index_15', caption: '執行条件', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                //執行条件（応答）
                return tjitems[record[15]]
            }},
            
            { field: 'index_7', caption: '約定価格', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                //価格
                return record[7];
            }},
            
            { field: 'index_16', caption: '決済相手<br>約定価格', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[16]
            }},
            
            { field: 'index_17', caption: '決済通貨', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[17]
            }},
            { field: 'index_18', caption: '為替差損益（実現損益）', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[18]
            }},
            
            { field: 'index_19', caption: 'スワップ', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[19]
            }},
             { field: 'index_20', caption: '円換算レート', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[20]
            }},           
            { field: 'index_1', caption: '注文ID（応答）', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[1]
            }},
            
            { field: 'index_2', caption: '銘柄', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[2]
            }},
            { field: 'index_3', caption: '建玉ID（応答）', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[3]
            }},
            { field: 'index_9', caption: '建玉金額（約定履歴）', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[9]
            }},
            { field: 'index_10', caption: '手数料', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[10]
            }},
            { field: 'index_11', caption: '現引金額', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[11]
            }},
            { field: 'index_12', caption: '受渡日', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[12]
            }},
            { field: 'index_13', caption: '備考', size: '10%',type: "text",hidden:true, render:function(record, index, column_index){
                return record[13]
            }},
"""


nameList = ["curid","tmid","mg","tgid","time","bs","xinki_type","price","amount","tgprice","tsr","ghprice","ukewatahi","biko","tm_method","require","ksyj_price","ks_paire","kawaseki_po","swap","ks_rate"]

# p = re.compile("[\w]*")
# p.match()

for x in range(len(nameList)):
    re.sub("index_"+ str(x) +"", pattern, nameList[x])
    re.sub("\brecord["+ str(x) +"]\b", pattern, "record[\""+ nameList[x] +"\"]")
    
print(pattern)

exit()

for x in range(len(nameList) -1, -1 ,-1):
    pattern = pattern.replace("index_" + str(x), nameList[x])
    pattern = pattern.replace("record["+str(x)+"]", "record[\""+ nameList[x] +"\"]" )
    
print(pattern)
    

exit()
wb = openpyxl.load_workbook('C:/Users/shikw/Desktop/ドキュメント/data.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
sheet = wb.get_sheet_by_name('Sheet1')
maxrow = sheet.max_row
maxcol = sheet.max_column


pattern = """{ field: 'index_{0}', caption: '{1}', size: '10%',type: "text",hidden:false, render:function(record, index, column_index){
                return record[{2}]
            }},
            """


for i in range(1,21):
    
    title = (sheet.cell(row=3,column=i).value or "")
    mystr = pattern.replace("{0}", str(i-1)).replace("{1}", title).replace("{2}", str(i-1))
    
    
    print(mystr, end="")

exit()
for i in range(1,maxrow):
    print("\n")
#     print("j" + i)
    for j in range(1,maxcol):
        print((sheet.cell(row=i,column=j).value or "" )+ "\t"  , end="") 
#         print("j" + str(j)+ "\t")
#         sys.stdout.write(".")
#         sys.stdout.flush()
        
        
    
