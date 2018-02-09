
# coding:utf-8

#https://www.zhihu.com/question/25566731

from _functools import reduce
from builtins import str
import json
import logging
import os
import re
import sqlite3
import sys
import unittest
import urllib

from lxml import etree
import lxml
import openpyxl
import openpyxl
from openpyxl.xml.constants import MAX_ROW
from pyquery import PyQuery as pq
import requests
from scrapper import Scrapper

from db import DB
import MeCab
import unicodedata
import threading

"""
    引用dll无法加载时，将dll拷贝到 python/site-packages下去
"""

# print(lxml.__file__)
try:
    import http.client as http_client
except ImportError:
    # Python 2
    import httplib as http_client
http_client.HTTPConnection.debuglevel = 1
 
logging.basicConfig()
logging.getLogger().setLevel(logging.DEBUG)
requests_log = logging.getLogger("requests.packages.urllib3")
requests_log.setLevel(logging.DEBUG)
requests_log.propagate = True




proxies = {
  'http': 'http://127.0.0.1:8888',
  'https': 'http://127.0.0.1:8888',
}

headers = {"Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
"Accept-Encoding":"gzip, deflate, br",
"Accept-Language":"en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7,ja;q=0.6",
"Cache-Control":"max-age=0",
"Connection":"keep-alive",
"Host":"dict.hjenglish.com",
"Upgrade-Insecure-Requests":"1",
"User-Agent":"Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36"}

# requests.get('https://stackoverflow.com/questions/10588644/how-can-i-see-the-entire-http-request-thats-being-sent-by-my-python-application')

# url = "https://dict.hjenglish.com/services/simpleExplain/jp_simpleExplain.ashx?type=jc&w=";
# resp.status_code == 200

# def doIt(x):
#     x["level"]= 3
#     return x
# dlist = [{"id":1},{"id":2}]
# print(list(map(doIt, dlist)))
# print(dlist)
# exit()


def getquerywordlist():
    wb = openpyxl.load_workbook('C:/Users/shikw/Desktop/samples/jpword.xlsx')

    level3Sheet= wb.get_sheet_by_name('Sheet1')
    level1Sheet = wb.get_sheet_by_name('Sheet2')
    querySheetArr = [level3Sheet, level1Sheet]
    resultArr = [None] * len(querySheetArr)
     
    for i in range(0,len(querySheetArr)):
        wordlist = []
        sheetvar = querySheetArr[i]
        maxrow = sheetvar.max_row
        maxcol = sheetvar.max_column
        
        for r in range(0,maxrow):
            rowdata = [None]*maxcol
            for c in range(0,maxcol):
                rowdata[c] = (sheetvar.cell(row=r+1,column=c+1).value or "")
            wordlist.append(rowdata)
                
        resultArr[i] = wordlist
    
    return resultArr

"""
#  import moduleName
#  dir(moduleName)
"""



def main():
    db = DB()
   
    dlist =  db.getwordtoquery()
    resultlist = []
    resultdatalist = []
    for row in dlist:
        id = row[0]
        word = row[1]
        querysuccess,info = loadfrominternel(word)
        resultlist.append(querysuccess)
        resultdatalist.extend(info)
    successlist = [ dlist[i][0] for i,x in enumerate(resultlist) if x == True]
    
    db.updatewordstatus(successlist,1)
    db.saveword(resultdatalist)
    return 
    

    
def loadfrominternel(word):
    
    result = []
    def doparse(index,node):
        
        d = pq(node)
        data = {}
        data["word"] = d(".word-info .word-text").text()
        data["jm"] = d(".word-info .pronounces span:eq(0)").text()
        data["roma"] = d(".word-info .pronounces span:eq(1)").text()
        data["sd"] = d(".word-info .pronounces span:eq(2)").text()
        data["wordtype"] = d(".simple span:eq(0)").text();
        data["simpleDefinition"] = d(".simple span:eq(1)").text();
        
        sens = {}
        
        def meaning_sens(index,node):
            d = pq(node)
            meaning = d("h3").text()
            
            sentences = []
            for item in d("ul li"):
                fromSentence = pq(item)("p:eq(0)").text()
                toSentence = pq(item)("p:eq(1)").text()
                senitem = {}
                senitem[fromSentence] = toSentence
                sentences.append(senitem)
            
            sens[meaning] = sentences
            
        data["sens"] = sens
        d(".word-details-item .detail-groups dd").each(meaning_sens)
        allResult.append(data)
        return
        
    url = "https://dict.hjenglish.com/jp/jc/" + word
    try:
         resp = doRequest(url, headers)
         if(resp.status_code == 200):
            content = resp.text
            d = pq(content)
            allResult = []
            d(".word-details-pane").each(doparse)
            result = allResult
    except BaseException as e:
        print(e)
    
    return len(result)>0 , result 

def loadfromnet(wordlist):
    
    wordinfolist = []
    replaceArr = ["content","IfHasScb","hjd_langs","WordId","FromLang","ToLang"]
    for i in range(0,len(wordlist)-1):
        word = wordlist[i]
        
        url = 'https://dict.hjenglish.com/services/simpleExplain/jp_simpleExplain.ashx?type=jc&w='+word
        resp = doRequest(url, headers)
        if(resp.status_code != 200):
            sys.exit("http request 异常" +resp.request)
        
        content = resp.text
        needed = content[content.find("{"): (content.rfind("}")+1) ]
        
        for c in range(0,len(replaceArr)):
            needed = needed.replace(replaceArr[c], "\"" + replaceArr[c] + "\"")
        dictjson = json.loads(needed)
        
        d = pq("<div>"+dictjson["content"] +"</div>")
        wordinfo = {}
        wordinfo["word"] = d("span.hjd_Green > font").html()
        wordinfo["roma"] = d("span:nth-child(2)").text()
        wordinfo["jm"] = d("span:nth-child(3)").text()
        wordinfo["sd"] = d("span:nth-child(4)").text()
        wordinfo["fyf"] = d("#hjd_wordcomment_1").attr("value")
        wordinfolist.append(wordinfo)
    return wordinfolist
    
    
def doRequest(url,headers):
     resp = requests.get(url,headers= headers)
     return resp


# db = DB()
# db.exesql("delete from jp_word")
# db.exesql("delete from jp_raw_word")
# 
# wordlistarr = getquerywordlist()
# 
# wordlistlevel3raw = wordlistarr[0]
# db.saveRawTableData(wordlistlevel3raw)
# wordlistlevel3 = list(map(lambda x:x[0], wordlistlevel3raw))[1:]
# commword = [x  for x in wordlistlevel3 if( x!="" and x.find("，")!=-1)]
# for i in range(0,len(commword)):
#     wordlistlevel3.extend(commword[i].split("，"))
# loadedwordinfolist1 = loadfromnet(list(set(wordlistlevel3)))
# for i in range(0,len(loadedwordinfolist1)):
#     loadedwordinfolist1[i]["level"] = 3
# db.saveWordList(loadedwordinfolist1)
# 
# 
# wordlistlevel1raw = wordlistarr[1]
# 
# db.saveRawTableData(wordlistlevel1raw)
# 
# wordlistlevel1 = list(map(lambda x:x[0], wordlistlevel1raw))[1:]
# 
# commword = [x  for x in wordlistlevel1 if( x!="" and x.find("，")!=-1)]
# for i in range(0,len(commword)):
#     wordlistlevel1.extend(commword[i].split("，"))
#     
# loadedwordinfolist1 = loadfromnet(list(set(wordlistlevel1)))
#     
# for i in range(0,len(loadedwordinfolist1)):
#     loadedwordinfolist1[i]["level"] = 1
# 
# db.saveWordList(loadedwordinfolist1)

def parseSentence(sens):
    mecab = MeCab.Tagger("-Ochasen")
    m = mecab.parseToNode(sens)
#     ignorewordtype = ["助詞"]
    wordlist = []
    while m:
        feature = m.feature
        feaarr = feature.split(",")
        originalword = feaarr[-3]
        wordlist.append(originalword)
        m = m.next
    return wordlist

def testmecab():
    mecab = MeCab.Tagger("-Ochasen")
    a1 = mecab.parse("太郎はこの本を二郎を見た女性に渡した。")
    
#     a2 = mecab.parse("使うほど、うるおいが満ちる")
    print(a1)

def is_japanese(string):
    for ch in string:
        name = unicodedata.name(ch) 
        if "CJK UNIFIED" in name \
        or "HIRAGANA" in name \
        or "KATAKANA" in name:
            return True
    return False


def starturlfetcherthread(starturl):
# レスん    https://qiita.com/konnyakmannan/items/2f0e3f00137db10f56a7#%E3%82%B5%E3%83%96%E3%82%AF%E3%83%A9%E3%82%B9%E3%82%92%E4%BD%9C%E3%82%8B
    threading.Thread(target=urlfetcherthread, name="url fetcher thread", args=(starturl))
    
def urlfetcherthread(starturl):
    
    
    return

def testscraper():
    scr = Scrapper()
    print("\n".join(scr.loadLink("https://qiita.com/kurohune538/items/55d2a9739b1f73363e56")))

if __name__ == '__main__':  # Script executed directly?
#     mlist = [1,2]
#     for x ,y in enumerate(mlist):
#         print(str(x) + " "+ str(y))
#     main()
#     testmecab()
#     resp = requests.get("https://qiita.com/kurohune538/items/55d2a9739b1f73363e56")
    testscraper()
